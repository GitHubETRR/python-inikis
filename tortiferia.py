from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment, numbers

#definiciones de cada objeto
objetos = {
    "TF": ("Tortafrita", 500),
    "MED": ("Medialuna", 600),
    "CAFE": ("Cafe", 250),
    "TOST": ("Tostado", 1000),
    "OFERTA": ("Oferta Cafe y 2 Medialunas", 1200),
    "TO(OFERTA)": ("Oferta Cafe y 2 Tostados", 1200)
}
metodos_pago = {
    "EF": "Efectivo",
    "MP": "Transferencia",
}

ventas = []
archivo_lista = open("ventas.txt",'r')
lista = archivo_lista.readlines()

#obtenemos los datos
for linea in lista:
    datos = linea.split()

    cant = int(datos[0])

    objeto = objetos[datos[1]][0] if objetos[datos[1]] else 0
    pago = metodos_pago[datos[2]] if metodos_pago[datos[2]] else 0
    precio = objetos[datos[1]][1]

    ventas.append((cant,objeto,pago,precio))

    #print("Se compro", objeto.lower(), "| Cantidad:", cant, "| Metodo de pago:", pago.lower())

archivo_lista.close()

#generamos el excel
excel = Workbook()
hoja = excel.active
hoja.title = "Ventas Tortiferia 12-9"

columnas = ("CANTIDAD","OBJETO","PAGO","PRECIO")
col_anchos = (11.7, 28.9, 13.3, 14.3) #respectivos a las columnas de arriba

#columnas
for col_index, col in enumerate(hoja.iter_cols(min_row=1, max_col=columnas.__len__())):
    for celda in col:
        celda.value = columnas[col_index]
        celda.font = Font(bold=True, color="FFFFFF")
        celda.alignment = Alignment(horizontal="center")
        celda.fill = PatternFill(fill_type="solid",start_color='000000')

        hoja.column_dimensions[celda.coordinate[0]].width = col_anchos[celda.col_idx-1]

#filas (items)
for fila_idx, fila in enumerate(hoja.iter_rows(min_row=2, max_col=columnas.__len__(), max_row=lista.__len__())):
    fila[0].value = ventas[fila_idx][0]
    fila[1].value = ventas[fila_idx][1]
    fila[2].value = ventas[fila_idx][2]
    fila[3].value = ventas[fila_idx][3]
    fila[3].number_format = '_-$ * #.##0,00_-;-$ * #.##0,00_-;_-$ * "-"??_-;_-@_-'

    for celda in fila:
        celda.alignment = Alignment(horizontal="center")

total_celda = hoja.cell(tuple(hoja.rows)[-1][0].row+1, 3, "TOTAL")
total_num_celda = hoja.cell(tuple(hoja.rows)[-1][0].row+1, 3, f"=SUM({tuple(hoja.rows)[3*2]}:{tuple(hoja.rows)[3*ventas.__len__()]})")

excel.save("ventas_12_9.xlsx")
